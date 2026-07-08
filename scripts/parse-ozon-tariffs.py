# -*- coding: utf-8 -*-
"""Parse Ozon tariff xlsx → JSON for unit economics calculator."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "ozon-tariffs"
OUT = ROOT / "src" / "lib" / "unit-economics" / "data" / "ozon"

VOLUME_RE = re.compile(
    r"(\d+(?:,\d+)?)\s*-\s*(\d+(?:,\d+)?)\s*л|до\s*(\d+(?:,\d+)?)\s*л|свыше\s*(\d+(?:,\d+)?)\s*л",
    re.I,
)


def parse_liters(s: str) -> tuple[float, float] | None:
    if not s:
        return None
    s = str(s).strip().replace("\u00a0", " ")
    m = re.match(r"(\d+(?:,\d+)?)\s*-\s*(\d+(?:,\d+)?)\s*л", s, re.I)
    if m:
        lo = float(m.group(1).replace(",", "."))
        hi = float(m.group(2).replace(",", "."))
        return lo, hi
    return None


def slugify(*parts: str) -> str:
    raw = "|".join(p.strip().lower() for p in parts if p and str(p).strip())
    raw = raw.replace("ё", "е")
    raw = re.sub(r"[^a-z0-9а-я|]+", "-", raw, flags=re.I)
    raw = re.sub(r"-+", "-", raw).strip("-")
    return raw[:180] or "unknown"


def pct(v) -> float:
    if v is None:
        return 0.0
    n = float(v)
    return round(n * 100, 3) if n <= 1 else round(n, 3)


def parse_categories(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    items = []
    seen = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[2]:
            continue
        parent = str(row[0] or "").strip()
        category = str(row[1] or "").strip()
        product_type = str(row[2] or "").strip()
        sid = slugify(parent, category, product_type)
        if sid in seen:
            sid = f"{sid}-{len(items)}"
        seen.add(sid)
        items.append(
            {
                "id": sid,
                "parentName": parent,
                "categoryName": category,
                "name": product_type,
                "label": f"{parent} · {category} · {product_type}",
                "commissionFbo": [pct(row[3]), pct(row[4]), pct(row[5])],
                "commissionFbs": [pct(row[9]), pct(row[10]), pct(row[11])],
            }
        )
    wb.close()
    return {
        "effectiveFrom": "2026-06-01",
        "sourceFile": path.name,
        "count": len(items),
        "items": items,
    }


def parse_logistics_sheet(ws, include_clusters: bool) -> tuple[list, list, list]:
    volume_bands: list[dict] = []
    band_index: dict[str, int] = {}
    ship_clusters: set[str] = set()
    delivery_clusters: set[str] = set()
    rates: list[dict] = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[1]:
            continue
        band_label = str(row[1]).strip()
        if band_label not in band_index:
            parsed = parse_liters(band_label)
            if not parsed:
                continue
            band_index[band_label] = len(volume_bands)
            volume_bands.append(
                {
                    "id": len(volume_bands),
                    "label": band_label,
                    "minLiters": parsed[0],
                    "maxLiters": parsed[1],
                }
            )
        bid = band_index[band_label]
        under300 = float(row[4 if include_clusters else 2] or 0)
        over300 = float(row[5 if include_clusters else 3] or 0)

        if include_clusters:
            ship = str(row[2] or "").strip()
            delivery = str(row[3] or "").strip()
            if not ship or not delivery:
                continue
            ship_clusters.add(ship)
            delivery_clusters.add(delivery)
            rates.append(
                {
                    "ship": ship,
                    "delivery": delivery,
                    "bandId": bid,
                    "under300": under300,
                    "over300": over300,
                }
            )
        else:
            rates.append({"bandId": bid, "under300": under300, "over300": over300})

    return (
        volume_bands,
        sorted(ship_clusters) if include_clusters else [],
        sorted(delivery_clusters) if include_clusters else [],
        rates,
    )


def parse_logistics(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rf = wb["Логистика РФ"]
    vb, ships, deliveries, matrix = parse_logistics_sheet(rf, True)
    default_ws = wb["Тарифы по умолчанию"]
    vb2, _, _, defaults = parse_logistics_sheet(default_ws, False)
    wb.close()
    if len(vb2) != len(vb):
        print("warn: default volume bands count differs", len(vb2), len(vb), file=sys.stderr)
    return {
        "effectiveFrom": "2026-05-01",
        "sourceFile": path.name,
        "volumeBands": vb,
        "shipClusters": ships,
        "deliveryClusters": deliveries,
        "matrix": matrix,
        "defaults": defaults,
    }


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    cat_path = next(SRC.glob("categories-*.xlsx"), None)
    log_path = next(SRC.glob("logistika-*.xlsx"), None)
    if not cat_path or not log_path:
        raise SystemExit(f"Missing xlsx in {SRC}")

    categories = parse_categories(cat_path)
    logistics = parse_logistics(log_path)

    marker = next((c for c in categories["items"] if c["name"] == "Маркер строительный"), None)
    msk = next((c for c in logistics["shipClusters"] if "Москва" in c), logistics["shipClusters"][0])

    meta = {
        "commissionsEffectiveFrom": categories["effectiveFrom"],
        "logisticsEffectiveFrom": logistics["effectiveFrom"],
        "commissionSource": categories["sourceFile"],
        "logisticsSource": logistics["sourceFile"],
        "categoryCount": categories["count"],
        "shipClusters": logistics["shipClusters"],
        "deliveryClusters": logistics["deliveryClusters"],
        "defaultCategoryId": marker["id"] if marker else categories["items"][0]["id"],
        "defaultShipCluster": msk,
        "defaultDeliveryCluster": msk,
        "acquiringPercentDefault": 1.5,
        "fbsProcessing": {
            "effectiveFrom": "2026-06-01",
            "pickupPerShipment": 20,
            "dropoffScTrust": 10,
            "dropoffScPiece": 20,
            "dropoffPvz": 30,
            "dropoffScTable": 20,
        },
        "fbsCargoUnit": {
            "effectiveFrom": "2026-04-16",
            "boxRub": 80,
            "palletRub": 175,
        },
    }

    (OUT / "categories.json").write_text(
        json.dumps(categories, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (OUT / "logistics.json").write_text(
        json.dumps(logistics, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    (OUT / "meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"categories: {categories['count']} -> {OUT / 'categories.json'}")
    print(f"logistics matrix rows: {len(logistics['matrix'])}")
    print(f"clusters ship={len(logistics['shipClusters'])} delivery={len(logistics['deliveryClusters'])}")
    if marker:
        print("marker sample:", marker["commissionFbo"], marker["commissionFbs"])
        print("default category:", marker["id"])
        print("default cluster:", msk)


if __name__ == "__main__":
    main()
