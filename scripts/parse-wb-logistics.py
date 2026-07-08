# -*- coding: utf-8 -*-
"""Parse WB logistics xlsx files → logistics.json for unit economics."""
from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RETURN = ROOT / "data" / "wb-tariffs" / "return-tariffs-2026-07-07.xlsx"
DEFAULT_WAREHOUSE = ROOT / "data" / "wb-tariffs" / "warehouse-coefficients-2026-07-07.xlsx"
DEFAULT_SNAPSHOTS = ROOT / "data" / "wb-tariffs" / "warehouse-tariff-snapshots.tsv"
OUT = ROOT / "src" / "lib" / "unit-economics" / "data" / "wb"

BASE_FIRST_LITER = 46.0
BASE_ADDITIONAL_LITER = 14.0
BASE_STORAGE_FIRST_LITER = 0.13
CUSTOMER_RETURN_RUB = 200.0
DEFAULT_MONOPALLET_ACCEPTANCE_PER_LITER = 4.17

LOCALIZATION_BRACKETS = [
    {"minShare": 95.0, "maxShare": 100.0, "localizationIndex": 0.50, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 90.0, "maxShare": 94.99, "localizationIndex": 0.60, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 85.0, "maxShare": 89.99, "localizationIndex": 0.70, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 80.0, "maxShare": 84.99, "localizationIndex": 0.80, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 75.0, "maxShare": 79.99, "localizationIndex": 0.90, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 70.0, "maxShare": 74.99, "localizationIndex": 1.00, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 65.0, "maxShare": 69.99, "localizationIndex": 1.00, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 60.0, "maxShare": 64.99, "localizationIndex": 1.00, "salesDistributionPercent": 0.0, "storageCoefPercent": 0.0},
    {"minShare": 55.0, "maxShare": 59.99, "localizationIndex": 1.05, "salesDistributionPercent": 2.0, "storageCoefPercent": 200.0},
    {"minShare": 50.0, "maxShare": 54.99, "localizationIndex": 1.10, "salesDistributionPercent": 2.05, "storageCoefPercent": 205.0},
    {"minShare": 45.0, "maxShare": 49.99, "localizationIndex": 1.20, "salesDistributionPercent": 2.05, "storageCoefPercent": 205.0},
    {"minShare": 40.0, "maxShare": 44.99, "localizationIndex": 1.30, "salesDistributionPercent": 2.10, "storageCoefPercent": 210.0},
    {"minShare": 35.0, "maxShare": 39.99, "localizationIndex": 1.40, "salesDistributionPercent": 2.10, "storageCoefPercent": 210.0},
    {"minShare": 30.0, "maxShare": 34.99, "localizationIndex": 1.50, "salesDistributionPercent": 2.15, "storageCoefPercent": 215.0},
    {"minShare": 25.0, "maxShare": 29.99, "localizationIndex": 1.55, "salesDistributionPercent": 2.20, "storageCoefPercent": 220.0},
    {"minShare": 20.0, "maxShare": 24.99, "localizationIndex": 1.60, "salesDistributionPercent": 2.25, "storageCoefPercent": 225.0},
    {"minShare": 15.0, "maxShare": 19.99, "localizationIndex": 1.70, "salesDistributionPercent": 2.30, "storageCoefPercent": 230.0},
    {"minShare": 10.0, "maxShare": 14.99, "localizationIndex": 1.75, "salesDistributionPercent": 2.35, "storageCoefPercent": 235.0},
    {"minShare": 5.0, "maxShare": 9.99, "localizationIndex": 1.80, "salesDistributionPercent": 2.45, "storageCoefPercent": 245.0},
    {"minShare": 0.0, "maxShare": 4.99, "localizationIndex": 2.00, "salesDistributionPercent": 2.50, "storageCoefPercent": 250.0},
]

SKIP_RETURN_NAMES = {"базовые тарифы"}


def normalize_key(name: str) -> str:
    raw = name.strip().lower().replace("ё", "е")
    for ch in "():,":
        raw = raw.replace(ch, " ")
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw


def slugify(name: str) -> str:
    translit = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ж": "zh", "з": "z",
        "и": "i", "й": "j", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p",
        "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "c", "ч": "ch",
        "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    raw = name.strip().lower().replace("ё", "е")
    latin = "".join(translit.get(ch, ch) for ch in raw)
    latin = re.sub(r"[^a-z0-9]+", "-", latin)
    latin = re.sub(r"-+", "-", latin).strip("-")
    return latin[:180] or "unknown"


def parse_money(raw) -> float | None:
    if raw is None:
        return None
    text = str(raw).strip().lower().replace("\u00a0", " ").replace(" ", "")
    if not text or text in {"непринимает", "не принимает", "-", "—"}:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_pct(raw) -> float | None:
    value = parse_money(raw)
    if value is None:
        return None
    return round(value, 4)


def parse_return_xlsx(path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    by_key: dict[str, dict] = {}
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if not label or normalize_key(label) in SKIP_RETURN_NAMES:
            continue
        key = normalize_key(label)
        by_key[key] = {
            "label": label,
            "returnToSellerPvzFirstRub": parse_money(row[1]),
            "returnToSellerPvzAdditionalRub": parse_money(row[2]),
            "returnToSellerCourierFirstRub": parse_money(row[3]),
            "returnToSellerCourierAdditionalRub": parse_money(row[4]),
            "returnUnclaimedRub": parse_money(row[5]),
            "returnCargoPvzFirstRub": parse_money(row[6]),
            "returnCargoPvzAdditionalRub": parse_money(row[7]),
            "returnCargoUnclaimedRub": parse_money(row[8]),
            "returnUnidentifiedPvzRub": parse_money(row[9]),
            "returnUnidentifiedUnclaimedRub": parse_money(row[10]),
        }
    return by_key


def parse_koroba_sheet(ws) -> dict[str, dict]:
    by_key: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        key = normalize_key(label)
        by_key[key] = {
            "label": label,
            "fbwCoefPercent": parse_pct(row[1]),
            "fbwFirstLiterRub": parse_money(row[2]),
            "fbwAdditionalLiterRub": parse_money(row[3]),
            "storageCoefPercent": parse_pct(row[4]),
            "storageFirstLiterDayRub": parse_money(row[5]),
            "storageAdditionalLiterDayRub": parse_money(row[6]),
            "fbsCoefPercent": parse_pct(row[7]),
            "fbsFirstLiterRub": parse_money(row[8]),
            "fbsAdditionalLiterRub": parse_money(row[9]),
        }
    return by_key


def parse_monopallet_sheet(ws) -> dict[str, dict]:
    by_key: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        key = normalize_key(label)
        by_key[key] = {
            "label": label,
            "monopalletFbwCoefPercent": parse_pct(row[1]),
            "monopalletFbwFirstLiterRub": parse_money(row[2]),
            "monopalletFbwAdditionalLiterRub": parse_money(row[3]),
            "monopalletStorageCoefPercent": parse_pct(row[4]),
            "monopalletStorageFirstPalletDayRub": parse_money(row[5]),
        }
    return by_key


def parse_warehouse_coefficients(path: Path) -> tuple[dict[str, dict], dict[str, dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    koroba = parse_koroba_sheet(wb[wb.sheetnames[0]])
    mono = parse_monopallet_sheet(wb[wb.sheetnames[1]]) if len(wb.sheetnames) > 1 else {}
    wb.close()
    return koroba, mono


def parse_snapshots_tsv(path: Path) -> dict[str, dict]:
    if not path.exists():
        return {}
    by_key: dict[str, dict] = {}
    with path.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh, delimiter="\t")
        for row in reader:
            label = (row.get("name") or "").strip()
            if not label:
                continue
            key = normalize_key(label)
            by_key[key] = {
                "label": label,
                "logisticsBoxScaled": parse_money(row.get("logistics_box")),
                "acceptanceBoxRub": parse_money(row.get("acceptance_box")),
                "acceptanceMonopalletRubPerLiter": parse_money(row.get("acceptance_monopallet")),
                "acceptsBoxSnapshot": row.get("accepts_box") == "1",
                "storageBoxDayMultiplier": parse_money(row.get("storage_box")),
            }
    return by_key


def pick_snapshot(snapshots_by_key: dict[str, dict], label: str) -> dict:
    return snapshots_by_key.get(normalize_key(label), {})


def pick_snapshot_alias(snapshots_by_key: dict[str, dict], label: str) -> dict:
    stripped = label.strip()
    lower = stripped.lower()
    alias = stripped[3:] if lower.startswith("сц ") else f"СЦ {stripped}"
    return snapshots_by_key.get(normalize_key(alias), {})


def merge_snapshot_fields(item: dict, snap: dict) -> None:
    if not snap:
        return
    for field in (
        "logisticsBoxScaled",
        "acceptanceBoxRub",
        "acceptanceMonopalletRubPerLiter",
        "storageBoxDayMultiplier",
        "acceptsBoxSnapshot",
    ):
        if item.get(field) is None and snap.get(field) is not None:
            item[field] = snap[field]
    if snap.get("acceptsBoxSnapshot"):
        item["acceptsBox"] = True
        item["profitCalculatorWarehouse"] = True


def merge_snapshot_acceptance_fields(item: dict, snap: dict) -> None:
    if not snap:
        return
    for field in ("acceptanceBoxRub", "acceptanceMonopalletRubPerLiter"):
        if item.get(field) is None and snap.get(field) is not None:
            item[field] = snap[field]


def derive_box_tariffs_from_snapshot(item: dict) -> None:
    logistics_box = item.get("logisticsBoxScaled")
    if logistics_box is None:
        return

    box_first = round(logistics_box * BASE_FIRST_LITER / 1000, 2)
    box_add = round(logistics_box * BASE_ADDITIONAL_LITER / 1000, 2)
    box_coef = round(box_first / BASE_FIRST_LITER * 100, 4)

    if item.get("fbwFirstLiterRub") is None and item.get("acceptsBoxSnapshot"):
        item["fbwFirstLiterRub"] = box_first
        item["fbwAdditionalLiterRub"] = box_add
        item["acceptsBox"] = True

    if item.get("fbwCoefPercent") is None:
        item["fbwCoefPercent"] = box_coef

    if item.get("fbsCoefPercent") is None and item.get("fbwCoefPercent") is not None:
        item["fbsCoefPercent"] = item["fbwCoefPercent"]

    if item.get("storageFirstLiterDayRub") is None and item.get("acceptsBox"):
        storage_coef = item.get("storageCoefPercent") or item.get("monopalletStorageCoefPercent") or 100
        item["storageCoefPercent"] = storage_coef
        rate = round(BASE_STORAGE_FIRST_LITER * storage_coef / 100, 4)
        item["storageFirstLiterDayRub"] = rate
        item["storageAdditionalLiterDayRub"] = rate

    if item.get("acceptanceMonopalletRubPerLiter") in (None, 0):
        item["acceptanceMonopalletRubPerLiter"] = DEFAULT_MONOPALLET_ACCEPTANCE_PER_LITER


def pick_label(*candidates: str | None) -> str:
    for candidate in candidates:
        if candidate and candidate.strip():
            return candidate.strip()
    return "unknown"


def build_warehouses(
    return_by_key: dict[str, dict],
    koroba_by_key: dict[str, dict],
    mono_by_key: dict[str, dict],
    snapshots_by_key: dict[str, dict],
) -> list[dict]:
    all_keys = sorted(set(return_by_key) | set(koroba_by_key) | set(mono_by_key) | set(snapshots_by_key))
    warehouses: list[dict] = []
    seen_ids: set[str] = set()

    for key in all_keys:
        ret = return_by_key.get(key, {})
        koroba = koroba_by_key.get(key, {})
        mono = mono_by_key.get(key, {})
        snap = pick_snapshot(snapshots_by_key, pick_label(ret.get("label"), koroba.get("label"), mono.get("label")))

        label = pick_label(ret.get("label"), koroba.get("label"), mono.get("label"), snap.get("label"))
        wid = slugify(label)
        if wid in seen_ids:
            wid = f"{wid}-{len(warehouses)}"
        seen_ids.add(wid)

        fbw_box_first = koroba.get("fbwFirstLiterRub")
        fbw_box_add = koroba.get("fbwAdditionalLiterRub")
        fbs_first = koroba.get("fbsFirstLiterRub")
        fbs_add = koroba.get("fbsAdditionalLiterRub")
        mono_first = mono.get("monopalletFbwFirstLiterRub")
        mono_add = mono.get("monopalletFbwAdditionalLiterRub")

        accepts_box = fbw_box_first is not None or fbs_first is not None
        accepts_monopallet = mono_first is not None

        if not accepts_box and not accepts_monopallet and not ret:
            continue

        item: dict = {
            "id": wid,
            "label": label,
            "acceptsBox": accepts_box,
            "acceptsMonopallet": accepts_monopallet,
            "customerReturnRub": CUSTOMER_RETURN_RUB,
            "fbwFirstLiterRub": fbw_box_first,
            "fbwAdditionalLiterRub": fbw_box_add,
            "fbsFirstLiterRub": fbs_first,
            "fbsAdditionalLiterRub": fbs_add,
            "fbwCoefPercent": koroba.get("fbwCoefPercent"),
            "fbsCoefPercent": koroba.get("fbsCoefPercent"),
            "storageCoefPercent": koroba.get("storageCoefPercent"),
            "storageFirstLiterDayRub": koroba.get("storageFirstLiterDayRub"),
            "storageAdditionalLiterDayRub": koroba.get("storageAdditionalLiterDayRub"),
            "monopalletFbwFirstLiterRub": mono_first,
            "monopalletFbwAdditionalLiterRub": mono_add,
            "monopalletFbwCoefPercent": mono.get("monopalletFbwCoefPercent"),
            "monopalletStorageCoefPercent": mono.get("monopalletStorageCoefPercent"),
            "monopalletStorageFirstPalletDayRub": mono.get("monopalletStorageFirstPalletDayRub"),
            "acceptanceBoxRub": snap.get("acceptanceBoxRub"),
            "acceptanceMonopalletRubPerLiter": snap.get("acceptanceMonopalletRubPerLiter"),
            "storageBoxDayMultiplier": snap.get("storageBoxDayMultiplier"),
            "hasKorobaStorageTariff": koroba.get("storageFirstLiterDayRub") is not None,
        }

        for k, v in ret.items():
            if k != "label" and v is not None:
                item[k] = v

        if fbw_box_first is not None:
            item["firstLiterRub"] = fbw_box_first
            item["additionalLiterRub"] = fbw_box_add
            item["logisticsCoef"] = round(fbw_box_first / BASE_FIRST_LITER, 6)
        elif fbs_first is not None:
            item["firstLiterRub"] = fbs_first
            item["additionalLiterRub"] = fbs_add
            item["logisticsCoef"] = round(fbs_first / BASE_FIRST_LITER, 6) if fbs_first else None
        elif mono_first is not None:
            item["firstLiterRub"] = mono_first
            item["additionalLiterRub"] = mono_add
            item["logisticsCoef"] = round(mono_first / BASE_FIRST_LITER, 6)

        warehouses.append(item)

    # Подмешиваем снапшоты (СЦ X → X) для складов с тарифами из xlsx.
    for item in warehouses:
        if item.get("fbwFirstLiterRub") is None and item.get("monopalletFbwFirstLiterRub") is None:
            continue
        snap = pick_snapshot(snapshots_by_key, item["label"])
        merge_snapshot_fields(item, snap)
        merge_snapshot_acceptance_fields(item, pick_snapshot_alias(snapshots_by_key, item["label"]))
        derive_box_tariffs_from_snapshot(item)
        if item.get("firstLiterRub") is None:
            if item.get("fbwFirstLiterRub") is not None:
                item["firstLiterRub"] = item["fbwFirstLiterRub"]
                item["additionalLiterRub"] = item["fbwAdditionalLiterRub"]
                item["logisticsCoef"] = round(item["fbwFirstLiterRub"] / BASE_FIRST_LITER, 6)
            elif item.get("monopalletFbwFirstLiterRub") is not None:
                item["firstLiterRub"] = item["monopalletFbwFirstLiterRub"]
                item["additionalLiterRub"] = item["monopalletFbwAdditionalLiterRub"]
                item["logisticsCoef"] = round(item["monopalletFbwFirstLiterRub"] / BASE_FIRST_LITER, 6)

    warehouses.sort(key=lambda w: w["label"].lower())
    return warehouses


def build_logistics(return_path: Path, warehouse_path: Path, snapshots_path: Path) -> dict:
    return_by_key = parse_return_xlsx(return_path)
    koroba_by_key: dict[str, dict] = {}
    mono_by_key: dict[str, dict] = {}
    if warehouse_path.exists():
        koroba_by_key, mono_by_key = parse_warehouse_coefficients(warehouse_path)
    snapshots_by_key = parse_snapshots_tsv(snapshots_path)
    warehouses = build_warehouses(return_by_key, koroba_by_key, mono_by_key, snapshots_by_key)

    source_files = [return_path.name]
    if warehouse_path.exists():
        source_files.append(warehouse_path.name)
    if snapshots_path.exists():
        source_files.append(snapshots_path.name)

    return {
        "effectiveFrom": "2026-07-07",
        "sourceFiles": source_files,
        "baseFirstLiterRub": BASE_FIRST_LITER,
        "baseAdditionalLiterRub": BASE_ADDITIONAL_LITER,
        "customerReturnRubDefault": CUSTOMER_RETURN_RUB,
        "localizationBrackets": LOCALIZATION_BRACKETS,
        "count": len(warehouses),
        "warehouses": warehouses,
    }


def main() -> None:
    return_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_RETURN
    warehouse_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_WAREHOUSE
    snapshots_path = Path(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_SNAPSHOTS

    if not return_path.exists():
        print(f"Return tariffs file not found: {return_path}", file=sys.stderr)
        sys.exit(1)

    OUT.mkdir(parents=True, exist_ok=True)
    logistics = build_logistics(return_path, warehouse_path, snapshots_path)

    (OUT / "logistics.json").write_text(
        json.dumps(logistics, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    meta_path = OUT / "meta.json"
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    meta["logisticsEffectiveFrom"] = logistics["effectiveFrom"]
    meta["logisticsSource"] = ", ".join(logistics["sourceFiles"])
    meta["defaultWarehouseId"] = "volgograd"
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    fbw = sum(1 for w in logistics["warehouses"] if w.get("fbwFirstLiterRub") is not None)
    mono = sum(1 for w in logistics["warehouses"] if w.get("acceptsMonopallet"))
    fbs = sum(1 for w in logistics["warehouses"] if w.get("fbsFirstLiterRub") is not None)
    print(
        f"warehouses={logistics['count']} fbw_box={fbw} monopallet={mono} fbs={fbs} "
        f"effectiveFrom={logistics['effectiveFrom']}"
    )
    print(f"written -> {OUT / 'logistics.json'}")


if __name__ == "__main__":
    main()
